""""
Copyright 2019 DevHyung
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import os
class ExcelDriver:
    def __init__(self, _fileName,_header):
        self.fileName = _fileName + ".xlsx"
        self.header = _header
        self.create_File()

    def create_File(self):
        if  os.path.exists(self.fileName):
            self.log('i','exist file, plz retry after remove the exist file')
        else:  # 새로만드는건
            book = Workbook()
            sheet = book.active
            sheet.title = 'default'
            sheet.append(self.header)
            # if modify cell width, write down
            #sheet.column_dimensions['A'].width = 40
            book.save(self.fileName)
    def set_col_width(self,widthDict):
        """
        :return: save excel that modified column width by user input
        """
        book = load_workbook(self.fileName)
        sheet = book.active
        for i in widthDict.items():
            sheet.column_dimensions[i[0]].width = i[1]
        book.save(self.fileName)

    def append_data(self,_data):
        '''
        :param _data: 1-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        sheet.append(_data)
        book.save(self.fileName)

    def append_data_list(self,_dataList):
        '''
        :param _dataList: 2-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        for data in _dataList:
            sheet.append(data)
        book.save(self.fileName)

    @staticmethod
    def log(tag, text):
        # Info tag
        if (tag.lower() == 'i'):
            print("[INFO] " + text)
        # Error tag
        elif (tag.lower() == 'e'):
            print("[ERROR] " + text)
        # Success tag
        elif (tag.lower() == 's'):
            print("[SUCCESS] " + text)
